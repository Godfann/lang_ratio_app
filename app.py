# -*- coding: utf-8 -*-
import os
import uuid
from pathlib import Path
from collections import Counter

import regex as re
from docx import Document as DocxDocument
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

ALLOWED_EXTENSIONS = {'docx', 'doc', 'txt', 'text'}

_job_store = {}


# ============================================================
# OPTIONAL TOKENIZERS
# ============================================================

try:
    from janome.tokenizer import Tokenizer as JanomeTokenizer
    JA_TOKENIZER = JanomeTokenizer()
except Exception:
    JA_TOKENIZER = None

try:
    import jieba
except Exception:
    jieba = None

try:
    from pythainlp.tokenize import word_tokenize as thai_word_tokenize
except Exception:
    thai_word_tokenize = None

try:
    from lingua import LanguageDetectorBuilder
    LINGUA_DETECTOR = LanguageDetectorBuilder.from_all_languages().build()
except Exception:
    LINGUA_DETECTOR = None


# ============================================================
# SUPPORTED LANGUAGE / LOCALE CODES
# ============================================================

SUPPORTED_LANGUAGE_CODES = {
    "EN": "English",
    "fr-CA": "Canadian French",
    "es-ES": "European Spanish",
    "fr-FR": "European French",
    "vi-VN": "Vietnamese",
    "it-IT": "Italian",
    "zh-HK": "Chinese Traditional Hong Kong",
    "pt-PT": "European Portuguese",
    "ar-AE": "Arabic UAE",
    "ar-SA": "Arabic Saudi Arabia",
    "pt-BR": "Brazilian Portuguese",
    "zh-CN": "Chinese Simplified China",
    "zh-TW": "Chinese Traditional Taiwan",
    "ja-JP": "Japanese",
    "ko-KR": "Korean",
    "th-TH": "Thai",
    "ru-RU": "Russian",
    "de-DE": "Standard German",
    "tr-TR": "Turkish",
    "FR": "French",
    "ES": "Spanish",
    "VI": "Vietnamese",
    "IT": "Italian",
    "ZH": "Chinese",
    "PT": "Portuguese",
    "AR": "Arabic",
    "JA": "Japanese",
    "KO": "Korean",
    "TH": "Thai",
    "RU": "Russian",
    "DE": "German",
    "TR": "Turkish",
    "UNKNOWN": "Unknown",
}

LANG_NAME_MAP = SUPPORTED_LANGUAGE_CODES.copy()


# ============================================================
# REGEX PATTERNS
# ============================================================

TAG_PATTERN = re.compile(
    r"\[([A-Za-z]{2,3}(?:[-_][A-Za-z]{2})?)\](.*?)\[/\1\]",
    re.DOTALL | re.IGNORECASE,
)

LEFTOVER_TAG_PATTERN = re.compile(
    r"\[/?[A-Za-z]{2,3}(?:[-_][A-Za-z]{2})?\]",
    re.IGNORECASE,
)

UNICODE_WORD_PATTERN = re.compile(
    r"[\p{L}\p{M}\p{N}]+(?:[''\-][\p{L}\p{M}\p{N}]+)*"
)

JA_PATTERN = re.compile(r"[\p{Script=Hiragana}\p{Script=Katakana}\p{Script=Han}]")
ZH_PATTERN = re.compile(r"[\p{Script=Han}]")
TH_PATTERN = re.compile(r"[\p{Script=Thai}]")


# ============================================================
# CLEANING AND NORMALIZATION
# ============================================================

def clean_text(text: str) -> str:
    return (
        str(text)
        .replace("﻿", "")
        .replace("```", "")
        .replace("``", "")
        .replace("`", "")
    )


def normalize_lang_code(lang: str) -> str:
    if not lang:
        return "UNKNOWN"

    raw = str(lang).strip().replace("_", "-")
    upper = raw.upper()

    alias_map = {
        "ENGLISH": "EN", "ENG": "EN", "EN": "EN", "EN-US": "EN",
        "EN-GB": "EN", "EN-IN": "EN",
        "FRENCH": "FR", "FR": "FR", "FRA": "FR", "FRE": "FR",
        "FR-CA": "fr-CA", "FR-FR": "fr-FR",
        "SPANISH": "ES", "ES": "ES", "SPA": "ES", "ES-ES": "es-ES",
        "VIETNAMESE": "VI", "VI": "VI", "VIE": "VI", "VI-VN": "vi-VN",
        "ITALIAN": "IT", "IT": "IT", "ITA": "IT", "IT-IT": "it-IT",
        "CHINESE": "ZH", "ZH": "ZH", "CN": "ZH", "CH": "ZH",
        "ZHO": "ZH", "CHI": "ZH",
        "ZH-HK": "zh-HK", "ZH-CN": "zh-CN", "ZH-TW": "zh-TW",
        "PORTUGUESE": "PT", "PT": "PT", "POR": "PT",
        "PT-PT": "pt-PT", "PT-BR": "pt-BR",
        "ARABIC": "AR", "AR": "AR", "ARA": "AR",
        "AR-AE": "ar-AE", "AR-SA": "ar-SA",
        "JAPANESE": "JA", "JA": "JA", "JP": "JA", "JPN": "JA", "JA-JP": "ja-JP",
        "KOREAN": "KO", "KO": "KO", "KOR": "KO", "KO-KR": "ko-KR",
        "THAI": "TH", "TH": "TH", "THA": "TH", "TH-TH": "th-TH",
        "RUSSIAN": "RU", "RU": "RU", "RUS": "RU", "RU-RU": "ru-RU",
        "GERMAN": "DE", "DE": "DE", "GER": "DE", "DEU": "DE", "DE-DE": "de-DE",
        "TURKISH": "TR", "TR": "TR", "TUR": "TR", "TR-TR": "tr-TR",
    }

    return alias_map.get(upper, raw)


def get_base_lang_code(lang: str) -> str:
    lang = normalize_lang_code(lang)
    if "-" in lang:
        return lang.split("-")[0].upper()
    return lang.upper()


# ============================================================
# WORD COUNT FUNCTIONS
# ============================================================

def count_japanese_words(text: str) -> int:
    if JA_TOKENIZER:
        count = 0
        for token in JA_TOKENIZER.tokenize(text):
            surface = token.surface.strip()
            pos = token.part_of_speech.split(",")[0]
            if not surface:
                continue
            if pos == "記号":
                continue
            count += 1
        return count
    return len(JA_PATTERN.findall(text))


def count_chinese_words(text: str) -> int:
    if jieba:
        tokens = [
            t.strip() for t in jieba.cut(text)
            if t.strip() and not re.fullmatch(r"\p{P}+", t.strip())
        ]
        return len(tokens)
    return len(ZH_PATTERN.findall(text))


def count_thai_words(text: str) -> int:
    if thai_word_tokenize:
        tokens = [
            t.strip() for t in thai_word_tokenize(text)
            if t.strip() and not re.fullmatch(r"\p{P}+", t.strip())
        ]
        return len(tokens)
    return len(TH_PATTERN.findall(text))


def count_general_words(text: str) -> int:
    return len(UNICODE_WORD_PATTERN.findall(text))


def count_words_by_language(lang: str, text: str) -> int:
    base = get_base_lang_code(lang)
    if base == "JA":
        return count_japanese_words(text)
    if base == "ZH":
        return count_chinese_words(text)
    if base == "TH":
        return count_thai_words(text)
    return count_general_words(text)


# ============================================================
# LANGUAGE DETECTION
# ============================================================

def detect_language_code(text: str) -> str:
    if not LINGUA_DETECTOR:
        return "UNKNOWN"
    detected = LINGUA_DETECTOR.detect_language_of(text)
    if not detected:
        return "UNKNOWN"
    lingua_to_code = {
        "ENGLISH": "EN", "FRENCH": "FR", "SPANISH": "ES",
        "VIETNAMESE": "VI", "ITALIAN": "IT", "CHINESE": "ZH",
        "PORTUGUESE": "PT", "ARABIC": "AR", "JAPANESE": "JA",
        "KOREAN": "KO", "THAI": "TH", "RUSSIAN": "RU",
        "GERMAN": "DE", "TURKISH": "TR",
    }
    return lingua_to_code.get(detected.name.upper(), "UNKNOWN")


# ============================================================
# INFER EXPECTED LANGUAGE CODES FROM FILE NAME
# ============================================================

def infer_expected_language_codes(file_name: str) -> list:
    text = file_name.lower()

    patterns = [
        ("fr-CA", [r"fr[-_]?ca", r"canadian\s*french"]),
        ("es-ES", [r"es[-_]?es", r"european\s*spanish"]),
        ("fr-FR", [r"fr[-_]?fr", r"european\s*french"]),
        ("vi-VN", [r"vi[-_]?vn", r"vietnamese"]),
        ("it-IT", [r"it[-_]?it", r"italian"]),
        ("zh-HK", [r"zh[-_]?hk", r"hong\s*kong"]),
        ("pt-PT", [r"pt[-_]?pt", r"european\s*portuguese"]),
        ("ar-AE", [r"ar[-_]?ae"]),
        ("ar-SA", [r"ar[-_]?sa", r"saudi\s*arabi"]),
        ("pt-BR", [r"pt[-_]?br", r"brazilian\s*portuguese"]),
        ("zh-CN", [r"zh[-_]?cn", r"simplified\s*chin"]),
        ("zh-TW", [r"zh[-_]?tw", r"traditional\s*taiwan", r"taiwan"]),
        ("ja-JP", [r"ja[-_]?jp", r"japanese"]),
        ("ko-KR", [r"ko[-_]?kr", r"korean"]),
        ("th-TH", [r"th[-_]?th", r"thai"]),
        ("ru-RU", [r"ru[-_]?ru", r"russian"]),
        ("de-DE", [r"de[-_]?de", r"(?:standard\s*)?german"]),
        ("tr-TR", [r"tr[-_]?tr", r"turkish"]),
        ("hi-IN", [r"hi[-_]?in", r"hindi"]),
    ]

    expected_codes = []
    for code, regex_list in patterns:
        for pattern in regex_list:
            if re.search(pattern, text, flags=re.IGNORECASE):
                expected_codes.append(code)
                break

    if re.search(r"\benglish\b|\beng\b|\ben\b", text, flags=re.IGNORECASE):
        expected_codes.append("EN")

    if expected_codes and "EN" not in expected_codes:
        expected_codes.append("EN")

    # Deduplicate while preserving order
    seen = set()
    final = []
    for code in expected_codes:
        if code not in seen:
            seen.add(code)
            final.append(code)

    return final


def align_results_to_expected_codes(results: list, expected_codes: list) -> list:
    if not expected_codes:
        return results

    expected_by_base = {}
    for code in expected_codes:
        base = get_base_lang_code(code)
        if "-" in code:
            expected_by_base[base] = code

    merged_counts = Counter()
    for row in results:
        code = normalize_lang_code(row["language_code"])
        base = get_base_lang_code(code)
        if "-" not in code and base in expected_by_base:
            code = expected_by_base[base]
        merged_counts[code] += row["word_count"]

    total_words = sum(merged_counts.values())
    aligned = []
    for code, count in merged_counts.most_common():
        ratio = (count / total_words * 100) if total_words else 0
        aligned.append({
            "language_code": code,
            "language_name": LANG_NAME_MAP.get(code, code),
            "word_count": count,
            "ratio_percent": round(ratio, 2),
        })
    return aligned


# ============================================================
# COUNT TAGGED TEXT AND UNTAGGED DIALOGUE
# ============================================================

def count_tagged_text(text: str) -> Counter:
    counts = Counter()
    for match in TAG_PATTERN.finditer(text):
        lang = normalize_lang_code(match.group(1))
        content = match.group(2).strip()
        if content:
            counts[lang] += count_words_by_language(lang, content)
    return counts


def count_untagged_dialogue(text: str) -> Counter:
    counts = Counter()
    text_without_tags = TAG_PATTERN.sub(" ", text)

    for line in text_without_tags.splitlines():
        line = line.strip()
        if not re.match(r"^(Agent|Cust|Customer)\s*:", line, flags=re.IGNORECASE):
            continue

        dialogue = re.sub(
            r"^(Agent|Cust|Customer)\s*:\s*", "", line, flags=re.IGNORECASE
        ).strip()
        dialogue = LEFTOVER_TAG_PATTERN.sub(" ", dialogue).strip()

        if not dialogue:
            continue

        detected = detect_language_code(dialogue)
        if detected == "UNKNOWN":
            if JA_PATTERN.search(dialogue):
                detected = "JA"
            elif TH_PATTERN.search(dialogue):
                detected = "TH"
            elif ZH_PATTERN.search(dialogue):
                detected = "ZH"

        counts[detected] += count_words_by_language(detected, dialogue)
    return counts


# ============================================================
# CALCULATE LANGUAGE RATIO FROM TEXT
# ============================================================

def calculate_language_ratio_from_text(text: str, include_untagged_dialogue: bool = True):
    text = clean_text(text)
    counts = count_tagged_text(text)
    if include_untagged_dialogue:
        counts += count_untagged_dialogue(text)

    total_words = sum(counts.values())
    results = []
    for lang, count in counts.most_common():
        ratio = (count / total_words * 100) if total_words else 0
        results.append({
            "language_code": lang,
            "language_name": LANG_NAME_MAP.get(lang, lang),
            "word_count": count,
            "ratio_percent": round(ratio, 2),
        })
    return results, total_words


# ============================================================
# READ TEXT FROM UPLOADED FILE (.docx or .txt)
# ============================================================

def read_file_text(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext in ('.docx', '.doc'):
        doc = DocxDocument(file_path)  # type: ignore[no-untyped-call]
        return '\n'.join(para.text for para in doc.paragraphs)  # type: ignore[attr-defined]
    return Path(file_path).read_text(encoding='utf-8', errors='ignore')


# ============================================================
# ANALYZE UPLOADED TEXT FILE  (entry point called by Flask)
# ============================================================

def analyze_text_file(file_path: str) -> dict:
    text = read_file_text(file_path)
    file_name = Path(file_path).stem

    expected_codes = infer_expected_language_codes(file_name)
    results, total_words = calculate_language_ratio_from_text(text)
    results = align_results_to_expected_codes(results, expected_codes)

    # Fill in any expected codes that had 0 words
    found_codes = {r["language_code"] for r in results}
    for code in expected_codes:
        if code not in found_codes:
            results.append({
                "language_code": code,
                "language_name": LANG_NAME_MAP.get(code, code),
                "word_count": 0,
                "ratio_percent": 0,
            })

    results = sorted(results, key=lambda x: x["word_count"], reverse=True)

    return {
        "file_name": file_name,
        "results": results,
        "total_words": total_words,
        "expected_codes": expected_codes,
    }


# ============================================================
# HELPERS
# ============================================================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# OUTPUT BUILDERS
# ============================================================

COLUMNS = [
    ('file name',               20),
    ('language code 1',         16),
    ('language code 2',         16),
    ('language code 1 words',   20),
    ('language code 2 words',   20),
    ('language code 1 Ratio %', 22),
    ('language code 2 Ratio %', 22),
    ('total words',             14),
    ('expected language codes', 24),
    ('status',                  10),
    ('error',                   30),
]

HEADER_FILL  = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F3F4FF")
THIN_BORDER  = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)


def result_to_row(file_name: str, results: list, total_words: int, expected_codes: list = None) -> dict:
    l1 = results[0] if len(results) > 0 else {}
    l2 = results[1] if len(results) > 1 else {}
    expected = ', '.join(expected_codes) if expected_codes else ', '.join(r['language_code'] for r in results)
    return {
        'file name':               file_name,
        'language code 1':         l1.get('language_code', ''),
        'language code 2':         l2.get('language_code', ''),
        'language code 1 words':   l1.get('word_count', 0),
        'language code 2 words':   l2.get('word_count', 0),
        'language code 1 Ratio %': l1.get('ratio_percent', 0),
        'language code 2 Ratio %': l2.get('ratio_percent', 0),
        'total words':             total_words,
        'expected language codes': expected,
        'status':                  'success',
        'error':                   '',
    }


def build_excel(file_name: str, results: list, total_words: int, out_path: str, expected_codes: list = None):
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = 'Summary'
    row_data = result_to_row(file_name, results, total_words, expected_codes)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws1.row_dimensions[1].height = 30

    col_names = [c[0] for c in COLUMNS]
    for col_idx, col_name in enumerate(col_names, start=1):
        val  = row_data.get(col_name, '')
        cell = ws1.cell(row=2, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = THIN_BORDER

    # Sheet 2: Detailed Breakdown
    ws2 = wb.create_sheet('Detailed Breakdown')
    detail_headers = ['Language Name', 'Language Code', 'Word Count', 'Ratio %']
    detail_widths  = [22, 18, 14, 12]

    for col_idx, (hdr, width) in enumerate(zip(detail_headers, detail_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=hdr)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[1].height = 28

    for r_idx, lang in enumerate(results, start=2):
        fill = ALT_ROW_FILL if r_idx % 2 == 0 else PatternFill()
        values = [lang['language_name'], lang['language_code'], lang['word_count'], lang['ratio_percent']]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = THIN_BORDER

    wb.save(out_path)


def build_pdf(file_name: str, results: list, total_words: int, out_path: str, expected_codes: list = None):
    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontSize=18, textColor=colors.HexColor('#1a1a2e'),
                                 spaceAfter=4)
    sub_style   = ParagraphStyle('Sub', parent=styles['Normal'],
                                 fontSize=10, textColor=colors.HexColor('#6b7280'),
                                 spaceAfter=18)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                   fontSize=12, textColor=colors.HexColor('#4f46e5'),
                                   spaceBefore=14, spaceAfter=8)

    story.append(Paragraph('Language Ratio Analysis', title_style))
    story.append(Paragraph(f'Script file: {file_name}', sub_style))

    # Summary table
    story.append(Paragraph('Summary', section_style))
    row_data = result_to_row(file_name, results, total_words, expected_codes)
    summary_data = [['Field', 'Value']]
    for col_name, _ in COLUMNS:
        val = row_data.get(col_name, '')
        summary_data.append([col_name.title(), str(val)])

    summary_table = Table(summary_data, colWidths=[6*cm, 10*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#4f46e5')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  10),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.HexColor('#f3f4ff'), colors.white]),
        ('FONTSIZE',      (0, 1), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)

    # Detailed breakdown
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph('Detailed Language Breakdown', section_style))

    detail_data = [['Language Name', 'Code', 'Words', 'Ratio %']]
    for lang in results:
        detail_data.append([
            lang['language_name'], lang['language_code'],
            str(lang['word_count']), f"{lang['ratio_percent']}%",
        ])

    col_w = [7*cm, 3.5*cm, 3.5*cm, 3*cm]
    detail_table = Table(detail_data, colWidths=col_w)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#4f46e5')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  10),
        ('ALIGN',         (0, 1), (-1, -1), 'CENTER'),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.HexColor('#f3f4ff'), colors.white]),
        ('FONTSIZE',      (0, 1), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    story.append(detail_table)

    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph(
        f'Total words analyzed: <b>{total_words:,}</b>',
        ParagraphStyle('footer', parent=styles['Normal'],
                       fontSize=9, textColor=colors.HexColor('#9ca3af')),
    ))

    doc.build(story)


# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/analyze', methods=['POST'])
def analyze():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Only .txt files are supported'}), 400

    job_id   = str(uuid.uuid4())
    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{job_id}_{filename}')
    file.save(file_path)

    try:
        analysis = analyze_text_file(file_path)
    except Exception as e:
        os.remove(file_path)
        return jsonify({'error': f'Analysis failed: {str(e)}'}), 500

    file_name      = analysis['file_name']
    results        = analysis['results']
    total_words    = analysis['total_words']
    expected_codes = analysis['expected_codes']

    excel_path = os.path.join(app.config['OUTPUT_FOLDER'], f'{job_id}.xlsx')
    pdf_path   = os.path.join(app.config['OUTPUT_FOLDER'], f'{job_id}.pdf')
    build_excel(file_name, results, total_words, excel_path, expected_codes)
    build_pdf(file_name, results, total_words, pdf_path, expected_codes)

    _job_store[job_id] = {
        'file_name':   file_name,
        'results':     results,
        'total_words': total_words,
        'excel_path':  excel_path,
        'pdf_path':    pdf_path,
    }

    os.remove(file_path)

    return jsonify({
        'job_id':      job_id,
        'file_name':   file_name,
        'results':     results,
        'total_words': total_words,
    })


@app.route('/download/<job_id>/<fmt>')
def download(job_id, fmt):
    job = _job_store.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404

    if fmt == 'excel':
        return send_file(job['excel_path'],
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f"{job['file_name']}_language_ratio.xlsx")
    elif fmt == 'pdf':
        return send_file(job['pdf_path'],
                         mimetype='application/pdf',
                         as_attachment=True,
                         download_name=f"{job['file_name']}_language_ratio.pdf")
    else:
        return jsonify({'error': 'Unknown format'}), 400


if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
    app.run(debug=True, port=5050)
